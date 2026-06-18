from __future__ import annotations

import calendar
import csv
import json
import re
import sys
from collections import Counter
from datetime import date
from decimal import Decimal, InvalidOperation
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


UPLOAD_HEADERS = [
    "분류",
    "관리유형",
    "납품일",
    "병원명",
    "담당자명",
    "연결그룹",
    "대표제품명",
    "세부모델/규격",
    "수량",
    "단가",
    "LOT번호",
    "유효기간만료일",
    "회계일자",
    "넣은시점",
    "품목코드",
    "기타작성칸",
]

EXCLUDE_HEADERS = [
    "제외사유",
    "원본행",
    "ERP일자",
    "판매처명",
    "품목군",
    "품목코드",
    "품명 및 규격",
    "수량",
    "단가",
    "합계",
    "적요",
    "담당자명",
    "시리얼no",
    "회계일자",
    "품목별적요",
]

CONSUMABLE_PRODUCTS = {"나비밴드", "나비픽스", "나잘스프린트", "나잘드레싱", "나빌룬", "유스타큐어 팁", "이관 시술 소모품", "기타 소모품"}


def clean(value) -> str:
    if value is None:
        return ""
    return str(value).replace("\xa0", " ").strip()


def parse_decimal(value) -> Decimal:
    if value is None or value == "":
        return Decimal("0")
    if isinstance(value, Decimal):
        return value
    if isinstance(value, (int, float)):
        return Decimal(str(value))
    text = str(value).strip().replace(",", "")
    if not text:
        return Decimal("0")
    try:
        return Decimal(text)
    except InvalidOperation:
        return Decimal("0")


def decimal_to_cell(value: Decimal):
    if value == value.to_integral():
        return int(value)
    return float(value)


def parse_erp_date(value) -> date | None:
    match = re.match(r"^(\d{2})/(\d{2})/(\d{2})", clean(value))
    if not match:
        return None
    yy, mm, dd = map(int, match.groups())
    try:
        return date(2000 + yy, mm, dd)
    except ValueError:
        return None


def add_years(value: date, years: int) -> date:
    try:
        return value.replace(year=value.year + years)
    except ValueError:
        return value.replace(year=value.year + years, day=28)


def find_serial_month(serial: str) -> tuple[int, int] | None:
    match = re.search(r"(20\d{2})(0[1-9]|1[0-2])", serial or "")
    if not match:
        return None
    return int(match.group(1)), int(match.group(2))


def expiry_for(product_group: str, delivery_date: date, serial: str) -> tuple[str, str]:
    if product_group not in CONSUMABLE_PRODUCTS:
        return "", "장비/비소모품 유효기간 미입력"

    serial_month = find_serial_month(serial)
    if serial_month:
        year, month = serial_month
        expiry_year = year + 3
        expiry_day = calendar.monthrange(expiry_year, month)[1]
        return date(expiry_year, month, expiry_day).isoformat(), f"시리얼 제조월 {year}-{month:02d} 기준 +3년"

    return add_years(delivery_date, 3).isoformat(), "시리얼 제조월 없음: 납품일 기준 +3년"


def contains_any(text: str, keys: list[str]) -> bool:
    upper = text.upper()
    return any(key.upper() in upper for key in keys)


def infer_mapping(item_group: str, item_code: str, item_name: str) -> tuple[str, str, str]:
    text = f"{item_group} {item_code} {item_name}"
    upper = text.upper()

    if contains_any(upper, ["1NB02", "NB-02", "NAVIBAND", "NAVI BAND"]):
        return "소모품", "나비밴드", "비강 소모품군"
    if contains_any(upper, ["DRESSING", "DFORCEP", "FORCEP", "33NASALD"]):
        return "소모품", "나잘드레싱", "비강 소모품군"
    if contains_any(upper, ["1SP-7070W", "MS-7070W", "7070"]):
        return "소모품", "나비픽스", "비강 소모품군"
    if contains_any(upper, ["33NASAL-SP", "NASAL SPLINT", "NASAL-SP", "33MDL-NSP"]):
        return "소모품", "나잘스프린트", "비강 소모품군"
    if contains_any(upper, ["32BALLOON", "NAVILLOON", "MG-BC-0601E"]):
        return "소모품", "나빌룬", "비강/이관 시술군"
    if contains_any(upper, ["1EUSTACURE", "MG-EUS"]):
        return "장비", "유스타큐어", "이관기능검사군"
    if contains_any(upper, ["33TIP-SONO", "33TIP-IMP", "TIP-SONO", "TIP-IMP"]):
        return "소모품", "유스타큐어 팁", "이관기능검사군"
    if contains_any(upper, ["EUSTACHIAN", "EU-CATH"]):
        return "소모품", "이관 시술 소모품", "이관기능검사군"
    if "장비" in item_group:
        return "장비", "기타 장비", ""
    return "소모품", "기타 소모품", ""


def nasal_splint_extra(item_name: str) -> str:
    if "NASAL SPLINT" not in item_name.upper():
        return ""
    models = re.findall(r"\[(MG-NS-[^\]]+)\]", item_name, flags=re.IGNORECASE)
    brackets = re.findall(r"\[([^\]]+)\]", item_name)
    model = models[0] if models else ""
    dimensions = brackets[-1] if brackets else ""
    if model and dimensions and model != dimensions:
        return f"모델상세:{model} {dimensions}"
    return f"모델상세:{model or dimensions or item_name}"


def compact_parts(parts: list[str]) -> str:
    return " / ".join(part for part in parts if part)


def fit_columns(worksheet, max_width: int = 52) -> None:
    for col_idx, column in enumerate(worksheet.columns, start=1):
        values = [clean(cell.value) for cell in list(column)[:200]]
        width = min(max(max([len(value) for value in values] + [8]) + 2, 10), max_width)
        worksheet.column_dimensions[get_column_letter(col_idx)].width = width


def style_worksheet(worksheet) -> None:
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    thin = Side(style="thin", color="D7DEE5")
    for cell in worksheet[1]:
        cell.font = Font(bold=True, color="17324D")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row in worksheet.iter_rows():
        for cell in row:
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    fit_columns(worksheet)


def write_xlsx(output_path: Path, payload: dict) -> None:
    workbook = Workbook()
    upload = workbook.active
    upload.title = "업로드용"
    upload.append(UPLOAD_HEADERS)
    for record in payload["included"]:
        upload.append([record.get(header, "") for header in UPLOAD_HEADERS])

    excluded = workbook.create_sheet("제외목록")
    excluded.append(EXCLUDE_HEADERS)
    for record in payload["excluded"]:
        excluded.append([record.get(header, "") for header in EXCLUDE_HEADERS])

    summary = workbook.create_sheet("요약")
    summary_rows = [
        ["항목", "값"],
        ["원본 파일", payload["summary"]["source"]],
        ["변환 기준 기간", payload["summary"]["period"]],
        ["업로드 포함 건수", payload["summary"]["included"]],
        ["제외 건수", payload["summary"]["excluded"]],
        ["소모품 포함 건수", payload["summary"]["assetCounts"].get("소모품", 0)],
        ["장비 포함 건수", payload["summary"]["assetCounts"].get("장비", 0)],
        ["변환 규칙", "0원/합계 0원, 수량 0 이하, 기간 외, 합계행은 업로드 제외"],
        ["유효기간 규칙", "소모품은 시리얼 제조월 기준 +3년, 시리얼 제조월이 없으면 납품일 기준 +3년"],
        [],
        ["대표제품명", "건수"],
    ]
    summary_rows.extend(payload["summary"]["productCounts"].items())
    summary_rows.extend([[], ["유효기간 계산 방식", "건수"]])
    summary_rows.extend(payload["summary"]["expiryMethodCounts"].items())
    summary_rows.extend([[], ["제외사유", "건수"]])
    summary_rows.extend(payload["summary"]["excludeReasonCounts"].items())
    summary_rows.extend([[], ["상위 병원/거래처", "건수"]])
    summary_rows.extend(payload["summary"]["topHospitals"].items())
    for row in summary_rows:
        summary.append(list(row))

    for worksheet in [upload, excluded, summary]:
        style_worksheet(worksheet)
    workbook.save(output_path)


def convert(source: Path, out_dir: Path) -> dict:
    out_dir.mkdir(parents=True, exist_ok=True)
    workbook = load_workbook(source, data_only=True, read_only=True)
    worksheet = workbook["판매현황"] if "판매현황" in workbook.sheetnames else workbook[workbook.sheetnames[0]]

    start = date(2020, 1, 1)
    end = date(2026, 4, 30)
    included = []
    excluded = []
    reasons = Counter()
    products = Counter()
    assets = Counter()
    hospitals = Counter()
    expiry_methods = Counter()

    for excel_row_idx, row in enumerate(worksheet.iter_rows(min_row=3, values_only=True), start=3):
        values = list(row) + [""] * 22
        erp_date_raw = clean(values[0])
        hospital = clean(values[3])
        item_group = clean(values[4])
        item_code = clean(values[5])
        item_name = clean(values[6])
        quantity = parse_decimal(values[7])
        unit_price = parse_decimal(values[8])
        total = parse_decimal(values[11])
        memo = clean(values[12])
        delivery_place = clean(values[13])
        manager = clean(values[14])
        serial = clean(values[15])
        accounting_raw = clean(values[18])
        accounting_date = parse_erp_date(accounting_raw)
        item_memo = clean(values[19])

        if not any([erp_date_raw, hospital, item_group, item_code, item_name, clean(values[7]), clean(values[8]), clean(values[11])]):
            continue

        parsed_date = parse_erp_date(erp_date_raw)
        reason = None
        if parsed_date is None:
            reason = "날짜 인식 불가/합계행"
        elif parsed_date < start or parsed_date > end:
            reason = "대상 기간 제외"
        elif unit_price == 0 or total == 0:
            reason = "0원 또는 합계 0원"
        elif quantity <= 0:
            reason = "수량 0 이하/반품 또는 보정"
        elif not hospital:
            reason = "병원명 없음"
        elif not item_name and not item_code:
            reason = "품목 정보 없음"

        if reason:
            reasons[reason] += 1
            excluded.append({
                "제외사유": reason,
                "원본행": excel_row_idx,
                "ERP일자": erp_date_raw,
                "판매처명": hospital,
                "품목군": item_group,
                "품목코드": item_code,
                "품명 및 규격": item_name,
                "수량": decimal_to_cell(quantity),
                "단가": decimal_to_cell(unit_price),
                "합계": decimal_to_cell(total),
                "적요": memo,
                "담당자명": manager,
                "시리얼no": serial,
                "회계일자": accounting_date.isoformat() if accounting_date else accounting_raw,
                "품목별적요": item_memo,
            })
            continue

        asset_type, product_group, relation_group = infer_mapping(item_group, item_code, item_name)
        expiry_date, expiry_method = expiry_for(product_group, parsed_date, serial)
        total_cell = decimal_to_cell(total)
        total_label = f"원본합계:{total_cell:,}" if isinstance(total_cell, int) else f"원본합계:{total_cell}"
        other = compact_parts([
            f"ERP일자번호:{erp_date_raw}",
            f"품목군:{item_group}" if item_group else "",
            nasal_splint_extra(item_name),
            f"유효기간계산:{expiry_method}",
            f"장비명:MG-EUS" if product_group == "유스타큐어" else "",
            f"회계일자번호:{accounting_raw}" if accounting_raw else "",
            f"적요:{memo}" if memo else "",
            f"납품장소:{delivery_place}" if delivery_place else "",
            f"품목별적요:{item_memo}" if item_memo else "",
            total_label,
        ])
        record = {
            "분류": asset_type,
            "관리유형": "신규납품",
            "납품일": parsed_date.isoformat(),
            "병원명": hospital,
            "담당자명": manager,
            "연결그룹": relation_group,
            "대표제품명": product_group,
            "세부모델/규격": item_name or item_code,
            "수량": decimal_to_cell(quantity),
            "단가": decimal_to_cell(unit_price),
            "LOT번호": serial,
            "유효기간만료일": expiry_date,
            "회계일자": accounting_date.isoformat() if accounting_date else "",
            "넣은시점": parsed_date.isoformat(),
            "품목코드": item_code,
            "기타작성칸": other,
        }
        included.append(record)
        products[product_group] += 1
        assets[asset_type] += 1
        hospitals[hospital] += 1
        expiry_methods[expiry_method] += 1

    upload_csv = out_dir / "ERP_bulk_upload_2020_20260430.csv"
    upload_xlsx = out_dir / "ERP_bulk_upload_2020_20260430.xlsx"
    excluded_csv = out_dir / "ERP_excluded_rows_2020_20260430.csv"
    json_path = out_dir / "ERP_conversion_payload_2020_20260430.json"
    summary_path = out_dir / "ERP_conversion_summary_2020_20260430.json"

    with upload_csv.open("w", newline="", encoding="utf-8-sig") as file:
        writer = csv.DictWriter(file, fieldnames=UPLOAD_HEADERS)
        writer.writeheader()
        writer.writerows(included)

    with excluded_csv.open("w", newline="", encoding="utf-8-sig") as file:
        writer = csv.DictWriter(file, fieldnames=EXCLUDE_HEADERS)
        writer.writeheader()
        writer.writerows(excluded)

    payload = {
        "uploadHeaders": UPLOAD_HEADERS,
        "excludeHeaders": EXCLUDE_HEADERS,
        "included": included,
        "excluded": excluded,
        "summary": {
            "source": str(source),
            "period": f"{start.isoformat()} ~ {end.isoformat()}",
            "included": len(included),
            "excluded": len(excluded),
            "assetCounts": dict(assets),
            "productCounts": dict(products.most_common()),
            "excludeReasonCounts": dict(reasons.most_common()),
            "expiryMethodCounts": dict(expiry_methods.most_common()),
            "topHospitals": dict(hospitals.most_common(20)),
        },
    }
    json_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    summary_path.write_text(json.dumps(payload["summary"], ensure_ascii=False, indent=2), encoding="utf-8")
    write_xlsx(upload_xlsx, payload)
    return payload["summary"]


if __name__ == "__main__":
    source_arg = Path(sys.argv[1]) if len(sys.argv) > 1 else Path(r"C:/Users/USER_001/Downloads/63NQ0VWZ67HCG34.xlsx")
    output_arg = Path(sys.argv[2]) if len(sys.argv) > 2 else Path("outputs/erp_import_2020_20260430")
    print(json.dumps(convert(source_arg, output_arg), ensure_ascii=False, indent=2))

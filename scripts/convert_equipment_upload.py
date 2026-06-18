from __future__ import annotations

import csv
import hashlib
import json
import re
import sys
from collections import Counter
from datetime import date, datetime, timezone
from decimal import Decimal, InvalidOperation
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


UPLOAD_HEADERS = [
    "데이터ID",
    "분류",
    "관리유형",
    "납품일",
    "병원명",
    "담당자명",
    "연결그룹",
    "대표제품명",
    "세부모델/규격",
    "수량",
    "단가",
    "LOT번호",
    "유효기간만료일",
    "회계일자",
    "넣은시점",
    "품목코드",
    "기타작성칸",
]

EXCLUDE_HEADERS = [
    "제외사유",
    "원본파일",
    "ERP일자",
    "판매처명",
    "품목코드",
    "품명 및 규격",
    "수량",
    "단가",
    "합계",
    "적요",
    "납품장소",
    "담당자명",
    "시리얼no",
    "회계전표일자-No.",
    "품목별적요",
]

PRODUCTS = {
    "eustachian": {
        "asset_type": "장비",
        "product_group": "이관기능검사기",
        "relation_group": "이관기능검사군",
        "keys": ["JK-05AD", "이관기능검사기", "1EUSTFM"],
    },
    "naviol": {
        "asset_type": "장비",
        "product_group": "NET-NAVIGATION (나비올)",
        "relation_group": "네비게이션 장비군",
        "keys": ["NET-NAVIGATION", "NAVIOL", "1VNAVIGATION"],
    },
}


def clean(value) -> str:
    if value is None:
        return ""
    return str(value).replace("\xa0", " ").strip()


def parse_decimal(value) -> Decimal:
    if value is None or value == "":
        return Decimal("0")
    if isinstance(value, Decimal):
        return value
    if isinstance(value, (int, float)):
        return Decimal(str(value))
    text = clean(value).replace(",", "")
    if not text:
        return Decimal("0")
    try:
        return Decimal(text)
    except InvalidOperation:
        return Decimal("0")


def decimal_to_cell(value: Decimal):
    if value == value.to_integral():
        return int(value)
    return float(value)


def parse_erp_date(value) -> date | None:
    match = re.match(r"^(\d{2})/(\d{2})/(\d{2})", clean(value))
    if not match:
        return None
    yy, mm, dd = map(int, match.groups())
    try:
        return date(2000 + yy, mm, dd)
    except ValueError:
        return None


def infer_product(item_code: str, item_name: str) -> dict | None:
    text = f"{item_code} {item_name}".upper()
    for product in PRODUCTS.values():
        if any(key.upper() in text for key in product["keys"]):
            return product
    return None


def compact_parts(parts: list[str]) -> str:
    return " / ".join(part for part in parts if part)


def deterministic_id(source_name: str, row_index: int, row: dict, product_group: str) -> str:
    raw = "|".join(
        [
            "erp-equipment",
            source_name,
            str(row_index),
            clean(row.get("일자")),
            clean(row.get("판매처명")),
            clean(row.get("품목코드")),
            clean(row.get("품명 및 규격")),
            clean(row.get("수량")),
            clean(row.get("단가")),
            clean(row.get("합 계")),
            clean(row.get("시리얼no")),
            product_group,
        ]
    )
    digest = hashlib.sha1(raw.encode("utf-8")).hexdigest()[:16]
    return f"erp-equipment-{digest}"


def row_to_excluded(source_name: str, row: dict, reason: str) -> dict:
    excluded = {"제외사유": reason, "원본파일": source_name}
    for header in EXCLUDE_HEADERS[2:]:
        excluded[header] = clean(row.get(header))
    return excluded


def parse_sheet(path: Path) -> tuple[list[dict], list[dict], Counter]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    header_index = None
    headers = []
    for index, row in enumerate(rows):
        if clean(row[0]) == "일자" and "품명 및 규격" in [clean(cell) for cell in row]:
            header_index = index
            headers = [clean(cell) for cell in row]
            break
    if header_index is None:
        raise ValueError(f"{path.name}: ERP 상세 헤더를 찾지 못했습니다.")

    included: list[dict] = []
    excluded: list[dict] = []
    reasons = Counter()

    for excel_row_index, values in enumerate(rows[header_index + 1 :], start=header_index + 2):
        row = {headers[i]: values[i] if i < len(values) else None for i in range(len(headers))}
        erp_date = parse_erp_date(row.get("일자"))
        if not erp_date:
            if any(clean(value) for value in values):
                reason = "ERP 납품일 없음 또는 합계/메모 행"
                excluded.append(row_to_excluded(path.name, row, reason))
                reasons[reason] += 1
            continue

        item_code = clean(row.get("품목코드"))
        item_name = clean(row.get("품명 및 규격"))
        product = infer_product(item_code, item_name)
        if not product:
            reason = "대상 장비 외 품목"
            excluded.append(row_to_excluded(path.name, row, reason))
            reasons[reason] += 1
            continue

        quantity = parse_decimal(row.get("수량"))
        unit_price = parse_decimal(row.get("단가"))
        total = parse_decimal(row.get("합 계")) or parse_decimal(row.get("공급가(부가세포함)"))
        if quantity <= 0:
            reason = "수량 0 이하 또는 회수 행"
            excluded.append(row_to_excluded(path.name, row, reason))
            reasons[reason] += 1
            continue
        if unit_price <= 0 or total <= 0:
            reason = "0원/무상/데모성 행"
            excluded.append(row_to_excluded(path.name, row, reason))
            reasons[reason] += 1
            continue

        accounting_date = parse_erp_date(row.get("회계전표일자-No."))
        memo = compact_parts(
            [
                f"ERP일자:{clean(row.get('일자'))}",
                f"합계:{decimal_to_cell(total):,}원",
                f"적요:{clean(row.get('적요'))}" if clean(row.get("적요")) else "",
                f"납품장소:{clean(row.get('납품장소'))}" if clean(row.get("납품장소")) else "",
                f"품목별적요:{clean(row.get('품목별적요'))}" if clean(row.get("품목별적요")) else "",
                f"회계전표:{clean(row.get('회계전표일자-No.'))}" if clean(row.get("회계전표일자-No.")) else "",
            ]
        )
        record = {
            "데이터ID": deterministic_id(path.name, excel_row_index, row, product["product_group"]),
            "분류": product["asset_type"],
            "관리유형": "신규납품",
            "납품일": erp_date.isoformat(),
            "병원명": clean(row.get("판매처명")) or clean(row.get("납품장소")),
            "담당자명": clean(row.get("담당자명")),
            "연결그룹": product["relation_group"],
            "대표제품명": product["product_group"],
            "세부모델/규격": item_name,
            "수량": decimal_to_cell(quantity),
            "단가": decimal_to_cell(unit_price),
            "LOT번호": clean(row.get("시리얼no")),
            "유효기간만료일": "",
            "회계일자": accounting_date.isoformat() if accounting_date else "",
            "넣은시점": date.today().isoformat(),
            "품목코드": item_code,
            "기타작성칸": memo,
        }
        included.append(record)

    return included, excluded, reasons


def to_supabase_record(record: dict) -> dict:
    now = datetime.now(timezone.utc).isoformat()
    return {
        "id": record["데이터ID"],
        "asset_type": record["분류"],
        "event_type": record["관리유형"],
        "delivery_date": record["납품일"] or None,
        "hospital": record["병원명"],
        "manager_name": record["담당자명"],
        "relation_group": record["연결그룹"],
        "product_group": record["대표제품명"],
        "product_detail": record["세부모델/규격"],
        "item_code": record["품목코드"],
        "quantity": record["수량"],
        "unit_price": record["단가"],
        "lot_number": record["LOT번호"],
        "expiry_date": None,
        "accounting_date": record["회계일자"] or None,
        "registered_at": record["넣은시점"] or None,
        "memo": record["기타작성칸"],
        "created_at": now,
        "updated_at": now,
    }


def fit_columns(worksheet, max_width: int = 54) -> None:
    for col_idx, column in enumerate(worksheet.columns, start=1):
        values = [clean(cell.value) for cell in list(column)[:250]]
        width = min(max(max([len(value) for value in values] + [8]) + 2, 10), max_width)
        worksheet.column_dimensions[get_column_letter(col_idx)].width = width


def style_worksheet(worksheet) -> None:
    header_fill = PatternFill("solid", fgColor="E6F0F7")
    thin = Side(style="thin", color="D7DEE5")
    for cell in worksheet[1]:
        cell.font = Font(bold=True, color="17324D")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row in worksheet.iter_rows():
        for cell in row:
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    fit_columns(worksheet)


def write_outputs(output_dir: Path, included: list[dict], excluded: list[dict], reasons: Counter, sources: list[str]) -> None:
    output_dir.mkdir(parents=True, exist_ok=True)
    xlsx_path = output_dir / "equipment_bulk_upload_2020_20260618.xlsx"
    csv_path = output_dir / "equipment_bulk_upload_2020_20260618.csv"
    excluded_csv_path = output_dir / "equipment_excluded_rows_2020_20260618.csv"
    json_path = output_dir / "equipment_supabase_records_2020_20260618.json"
    summary_path = output_dir / "equipment_conversion_summary_2020_20260618.json"

    workbook = Workbook()
    upload = workbook.active
    upload.title = "업로드용"
    upload.append(UPLOAD_HEADERS)
    for record in included:
        upload.append([record.get(header, "") for header in UPLOAD_HEADERS])

    excluded_sheet = workbook.create_sheet("제외목록")
    excluded_sheet.append(EXCLUDE_HEADERS)
    for record in excluded:
        excluded_sheet.append([record.get(header, "") for header in EXCLUDE_HEADERS])

    summary = workbook.create_sheet("요약")
    product_counts = Counter(record["대표제품명"] for record in included)
    summary_rows = [
        ["항목", "값"],
        ["원본 파일", ", ".join(sources)],
        ["변환 기준", "장비 / 금액·합계 0원 제외 / 수량 0 이하·회수 제외 / 합계행 제외 / 대상 장비만 포함"],
        ["업로드 포함 건수", len(included)],
        ["제외 건수", len(excluded)],
        ["회계일자 포함 건수", sum(1 for record in included if record["회계일자"])],
        [],
        ["대표제품명", "건수"],
    ]
    for product_name, count in product_counts.most_common():
        summary_rows.append([product_name, count])
    summary_rows.append([])
    summary_rows.append(["제외사유", "건수"])
    for reason, count in reasons.most_common():
        summary_rows.append([reason, count])
    for row in summary_rows:
        summary.append(row)

    for sheet in workbook.worksheets:
        style_worksheet(sheet)
    workbook.save(xlsx_path)

    with csv_path.open("w", newline="", encoding="utf-8-sig") as file:
        writer = csv.DictWriter(file, fieldnames=UPLOAD_HEADERS)
        writer.writeheader()
        writer.writerows(included)

    with excluded_csv_path.open("w", newline="", encoding="utf-8-sig") as file:
        writer = csv.DictWriter(file, fieldnames=EXCLUDE_HEADERS)
        writer.writeheader()
        writer.writerows(excluded)

    supabase_records = [to_supabase_record(record) for record in included]
    json_path.write_text(json.dumps(supabase_records, ensure_ascii=False, indent=2), encoding="utf-8")
    summary_payload = {
        "sources": sources,
        "included": len(included),
        "excluded": len(excluded),
        "accountingDateIncluded": sum(1 for record in included if record["회계일자"]),
        "productCounts": dict(product_counts),
        "excludeReasons": dict(reasons),
        "xlsx": str(xlsx_path),
        "csv": str(csv_path),
        "excludedCsv": str(excluded_csv_path),
        "supabaseJson": str(json_path),
    }
    summary_path.write_text(json.dumps(summary_payload, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps(summary_payload, ensure_ascii=False, indent=2))


def main() -> int:
    if len(sys.argv) < 3:
        print("Usage: convert_equipment_upload.py <output_dir> <source.xlsx> [source.xlsx ...]", file=sys.stderr)
        return 2
    output_dir = Path(sys.argv[1])
    source_paths = [Path(arg) for arg in sys.argv[2:]]
    included: list[dict] = []
    excluded: list[dict] = []
    reasons = Counter()

    for path in source_paths:
        if not path.exists():
            print(f"missing: {path}", file=sys.stderr)
            continue
        source_included, source_excluded, source_reasons = parse_sheet(path)
        included.extend(source_included)
        excluded.extend(source_excluded)
        reasons.update(source_reasons)

    included.sort(key=lambda record: (record["납품일"], record["병원명"], record["대표제품명"]))
    write_outputs(output_dir, included, excluded, reasons, [path.name for path in source_paths if path.exists()])
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

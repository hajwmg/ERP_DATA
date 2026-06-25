from __future__ import annotations

import calendar
import csv
import hashlib
import json
import re
import sys
from collections import Counter
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


UPLOAD_HEADERS = [
    "데이터ID",
    "분류",
    "관리유형",
    "납품일",
    "병원명",
    "담당자명",
    "연결그룹",
    "대표제품명",
    "세부모델/규격",
    "수량",
    "단가",
    "LOT번호",
    "유효기간만료일",
    "회계일자",
    "넣은시점",
    "품목코드",
    "기타작성칸",
]

EXCLUDE_HEADERS = [
    "제외사유",
    "원본파일",
    "거래명세서번호",
    "거래명세서일",
    "거래처",
    "유통구조",
    "납품장소",
    "품명",
    "품번",
    "수량",
    "판매단가",
    "판매금액계",
    "Lot/특이사항",
]

PRODUCT_MAPPINGS = [
    ("소모품", "나비밴드", "비강 소모품군", ["1NB02", "NB-02", "NAVIBAND", "NAVI BAND"]),
    ("소모품", "나비픽스", "비강 소모품군", ["1SP-7070W", "MS-7070W", "7070"]),
    ("소모품", "나잘스프린트", "비강 소모품군", ["33NASAL-SP", "NASAL SPLINT", "NASAL-SP", "1NASAL-SP", "33MDL-NSP"]),
    ("소모품", "나잘드레싱", "비강 소모품군", ["DRESSING", "DFORCEP", "FORCEP", "33NASALD", "1NASALD", "MG-NP"]),
    ("소모품", "나빌룬", "비강/이관 시술군", ["32BALLOON", "NAVILLOON", "MG-BC-0601E", "1BALLOON"]),
    ("소모품", "유스타큐어 팁", "이관기능검사군", ["33TIP-SONO", "33TIP-IMP", "TIP-SONO", "TIP-IMP", "33TIP-EUSTACURE"]),
    ("장비", "이관기능검사기", "이관기능검사군", ["JK-05AD", "1EUSTFM", "이관기능검사기"]),
    ("장비", "NET-NAVIGATION (나비올)", "네비게이션 장비군", ["NET-NAVIGATION", "NAVIOL", "1VNAVIGATION"]),
]


def clean(value) -> str:
    if value is None:
        return ""
    return str(value).replace("\xa0", " ").strip()


def parse_decimal(value) -> Decimal:
    if value is None or value == "":
        return Decimal("0")
    if isinstance(value, Decimal):
        return value
    if isinstance(value, (int, float)):
        return Decimal(str(value))
    text = clean(value).replace(",", "")
    if not text:
        return Decimal("0")
    try:
        return Decimal(text)
    except InvalidOperation:
        return Decimal("0")


def decimal_to_cell(value: Decimal):
    if value == value.to_integral():
        return int(value)
    return float(value)


def parse_date(value) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = clean(value).replace(".", "-").replace("/", "-")
    match = re.match(r"^(\d{4})-(\d{1,2})-(\d{1,2})", text)
    if not match:
        return None
    year, month, day = map(int, match.groups())
    try:
        return date(year, month, day)
    except ValueError:
        return None


def add_years(value: date, years: int) -> date:
    try:
        return value.replace(year=value.year + years)
    except ValueError:
        return value.replace(year=value.year + years, day=28)


def find_serial_month(serial: str) -> tuple[int, int] | None:
    text = clean(serial)
    match = re.search(r"(20\d{2})(0[1-9]|1[0-2])", text)
    if match:
        return int(match.group(1)), int(match.group(2))
    short_match = re.search(r"\((\d{2})[./-](0[1-9]|1[0-2])\)", text)
    if short_match:
        return 2000 + int(short_match.group(1)), int(short_match.group(2))
    return None


def expiry_for(asset_type: str, delivery_date: date, lot_text: str) -> str:
    if asset_type != "소모품":
        return ""
    serial_month = find_serial_month(lot_text)
    if serial_month:
        year, month = serial_month
        expiry_year = year + 3
        expiry_day = calendar.monthrange(expiry_year, month)[1]
        return date(expiry_year, month, expiry_day).isoformat()
    return add_years(delivery_date, 3).isoformat()


def contains_any(text: str, keys: list[str]) -> bool:
    upper = text.upper()
    return any(key.upper() in upper for key in keys)


def infer_product(item_code: str, item_name: str):
    text = f"{item_code} {item_name}"
    for asset_type, product_group, relation_group, keys in PRODUCT_MAPPINGS:
        if contains_any(text, keys):
            return asset_type, product_group, relation_group
    return None


def compact_parts(parts: list[str]) -> str:
    return " / ".join(part for part in parts if part)


def row_value(values, index: int):
    return values[index] if len(values) > index else None


def deterministic_id(source_name: str, values) -> str:
    raw = "|".join(
        [
            "statement-import",
            clean(row_value(values, 2)),
            clean(row_value(values, 3)),
            clean(row_value(values, 9)),
            clean(row_value(values, 20)),
            clean(row_value(values, 19)),
            clean(row_value(values, 24)),
            clean(row_value(values, 28)),
            clean(row_value(values, 31)),
            clean(row_value(values, 37)),
            clean(row_value(values, 40)),
        ]
    )
    digest = hashlib.sha1(raw.encode("utf-8")).hexdigest()[:16]
    return f"statement-{digest}"


def unique_key(values) -> tuple[str, ...]:
    return (
        clean(row_value(values, 2)),
        clean(row_value(values, 3)),
        clean(row_value(values, 9)),
        clean(row_value(values, 20)),
        clean(row_value(values, 19)),
        clean(row_value(values, 24)),
        clean(row_value(values, 28)),
        clean(row_value(values, 31)),
        clean(row_value(values, 37)),
        clean(row_value(values, 40)),
    )


def excluded_row(source_name: str, values, reason: str) -> dict:
    lot = compact_parts([clean(row_value(values, 37)), clean(row_value(values, 40))])
    return {
        "제외사유": reason,
        "원본파일": source_name,
        "거래명세서번호": clean(row_value(values, 2)),
        "거래명세서일": clean(row_value(values, 3)),
        "거래처": clean(row_value(values, 9)),
        "유통구조": clean(row_value(values, 10)),
        "납품장소": clean(row_value(values, 13)),
        "품명": clean(row_value(values, 19)),
        "품번": clean(row_value(values, 20)),
        "수량": clean(row_value(values, 24)),
        "판매단가": clean(row_value(values, 28)),
        "판매금액계": clean(row_value(values, 31)),
        "Lot/특이사항": lot,
    }


def parse_files(source_paths: list[Path]):
    included: list[dict] = []
    excluded: list[dict] = []
    reasons = Counter()
    seen = set()

    for path in source_paths:
        workbook = load_workbook(path, read_only=True, data_only=True)
        worksheet = workbook.active
        for values in worksheet.iter_rows(min_row=3, values_only=True):
            if clean(row_value(values, 1)) == "TOTAL":
                continue
            statement_no = clean(row_value(values, 2))
            if not statement_no:
                continue
            key = unique_key(values)
            if key in seen:
                reason = "중복 행"
                excluded.append(excluded_row(path.name, values, reason))
                reasons[reason] += 1
                continue
            seen.add(key)

            item_name = clean(row_value(values, 19))
            item_code = clean(row_value(values, 20))
            mapping = infer_product(item_code, item_name)
            if not mapping:
                reason = "관리 대상 외 품목"
                excluded.append(excluded_row(path.name, values, reason))
                reasons[reason] += 1
                continue

            quantity = parse_decimal(row_value(values, 24))
            unit_price = parse_decimal(row_value(values, 28))
            total_amount = parse_decimal(row_value(values, 31))
            if quantity <= 0:
                reason = "수량 0 이하"
                excluded.append(excluded_row(path.name, values, reason))
                reasons[reason] += 1
                continue
            if unit_price <= 0 or total_amount <= 0:
                reason = "0원/무상/샘플성 행"
                excluded.append(excluded_row(path.name, values, reason))
                reasons[reason] += 1
                continue

            delivery_date = parse_date(row_value(values, 3))
            if not delivery_date:
                reason = "거래명세서일 없음"
                excluded.append(excluded_row(path.name, values, reason))
                reasons[reason] += 1
                continue

            asset_type, product_group, relation_group = mapping
            lot_text = compact_parts([clean(row_value(values, 37)), clean(row_value(values, 40))])
            tax_status = clean(row_value(values, 38))
            sales_timing = clean(row_value(values, 42))
            accounting_date = delivery_date.isoformat() if tax_status == "완료" or sales_timing else ""
            product_detail = compact_parts([item_name, clean(row_value(values, 21))])
            memo = compact_parts(
                [
                    f"거래명세서:{statement_no}",
                    f"청구처:{clean(row_value(values, 8))}" if clean(row_value(values, 8)) else "",
                    f"유통구조:{clean(row_value(values, 10))}" if clean(row_value(values, 10)) else "",
                    f"납품장소:{clean(row_value(values, 13))}" if clean(row_value(values, 13)) else "",
                    f"출고구분:{clean(row_value(values, 5))}" if clean(row_value(values, 5)) else "",
                    f"세금계산서:{tax_status}" if tax_status else "",
                    f"매출시점:{sales_timing}" if sales_timing else "",
                    f"판매금액계:{decimal_to_cell(total_amount):,}원",
                ]
            )
            included.append(
                {
                    "데이터ID": deterministic_id(path.name, values),
                    "분류": asset_type,
                    "관리유형": "신규납품",
                    "납품일": delivery_date.isoformat(),
                    "병원명": clean(row_value(values, 9)) or clean(row_value(values, 8)),
                    "담당자명": clean(row_value(values, 7)),
                    "연결그룹": relation_group,
                    "대표제품명": product_group,
                    "세부모델/규격": product_detail,
                    "수량": decimal_to_cell(quantity),
                    "단가": decimal_to_cell(unit_price),
                    "LOT번호": lot_text,
                    "유효기간만료일": expiry_for(asset_type, delivery_date, lot_text),
                    "회계일자": accounting_date,
                    "넣은시점": date.today().isoformat(),
                    "품목코드": item_code,
                    "기타작성칸": memo,
                }
            )

    included.sort(key=lambda record: (record["납품일"], record["병원명"], record["대표제품명"], record["품목코드"]))
    return included, excluded, reasons


def fit_columns(worksheet, max_width: int = 58) -> None:
    for col_idx, column in enumerate(worksheet.columns, start=1):
        values = [clean(cell.value) for cell in list(column)[:250]]
        width = min(max(max([len(value) for value in values] + [8]) + 2, 10), max_width)
        worksheet.column_dimensions[get_column_letter(col_idx)].width = width


def style_worksheet(worksheet) -> None:
    header_fill = PatternFill("solid", fgColor="E6F0F7")
    thin = Side(style="thin", color="D7DEE5")
    for cell in worksheet[1]:
        cell.font = Font(bold=True, color="17324D")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row in worksheet.iter_rows():
        for cell in row:
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    fit_columns(worksheet)


def write_outputs(output_dir: Path, included: list[dict], excluded: list[dict], reasons: Counter, sources: list[Path]) -> None:
    output_dir.mkdir(parents=True, exist_ok=True)
    xlsx_path = output_dir / "statement_bulk_upload_20260506_20260605.xlsx"
    csv_path = output_dir / "statement_bulk_upload_20260506_20260605.csv"
    excluded_csv_path = output_dir / "statement_excluded_rows_20260506_20260605.csv"
    summary_path = output_dir / "statement_conversion_summary_20260506_20260605.json"

    workbook = Workbook()
    upload = workbook.active
    upload.title = "업로드용"
    upload.append(UPLOAD_HEADERS)
    for record in included:
        upload.append([record.get(header, "") for header in UPLOAD_HEADERS])

    excluded_sheet = workbook.create_sheet("제외목록")
    excluded_sheet.append(EXCLUDE_HEADERS)
    for record in excluded:
        excluded_sheet.append([record.get(header, "") for header in EXCLUDE_HEADERS])

    summary_sheet = workbook.create_sheet("요약")
    product_counts = Counter(record["대표제품명"] for record in included)
    summary_rows = [
        ["항목", "값"],
        ["원본 파일", ", ".join(path.name for path in sources)],
        ["변환 기준", "기존 관리 제품만 포함 / 0원·무상·수량 0 이하·중복·미매핑 품목 제외"],
        ["업로드 포함 건수", len(included)],
        ["제외 건수", len(excluded)],
        [],
        ["대표제품명", "건수"],
    ]
    for product_name, count in product_counts.most_common():
        summary_rows.append([product_name, count])
    summary_rows.append([])
    summary_rows.append(["제외사유", "건수"])
    for reason, count in reasons.most_common():
        summary_rows.append([reason, count])
    for row in summary_rows:
        summary_sheet.append(row)

    for sheet in workbook.worksheets:
        style_worksheet(sheet)
    workbook.save(xlsx_path)

    with csv_path.open("w", newline="", encoding="utf-8-sig") as file:
        writer = csv.DictWriter(file, fieldnames=UPLOAD_HEADERS)
        writer.writeheader()
        writer.writerows(included)

    with excluded_csv_path.open("w", newline="", encoding="utf-8-sig") as file:
        writer = csv.DictWriter(file, fieldnames=EXCLUDE_HEADERS)
        writer.writeheader()
        writer.writerows(excluded)

    payload = {
        "sources": [path.name for path in sources],
        "included": len(included),
        "excluded": len(excluded),
        "productCounts": dict(product_counts),
        "excludeReasons": dict(reasons),
        "xlsx": str(xlsx_path),
        "csv": str(csv_path),
        "excludedCsv": str(excluded_csv_path),
    }
    summary_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps(payload, ensure_ascii=False, indent=2))


def main() -> int:
    if len(sys.argv) < 3:
        print("Usage: convert_statement_upload.py <output_dir> <source.xlsx> [source.xlsx ...]", file=sys.stderr)
        return 2
    output_dir = Path(sys.argv[1])
    source_paths = [Path(arg) for arg in sys.argv[2:] if Path(arg).exists()]
    included, excluded, reasons = parse_files(source_paths)
    write_outputs(output_dir, included, excluded, reasons, source_paths)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
